from moviepy.editor import VideoFileClip
import openpyxl
from openpyxl.drawing.image import Image as OImage
import os
from pathlib import Path
from PIL import Image
import math
from flask import redirect, render_template, session
from functools import wraps



def apology(message, code=400):
    """Render message as an apology to user."""

    def escape(s):
        """
        Escape special characters.

        https://github.com/jacebrowning/memegen#special-characters
        """
        for old, new in [
            ("-", "--"),
            (" ", "-"),
            ("_", "__"),
            ("?", "~q"),
            ("%", "~p"),
            ("#", "~h"),
            ("/", "~s"),
            ('"', "''"),
        ]:
            s = s.replace(old, new)
        return s

    return render_template("apology.html", top=code, bottom=escape(message)), code


def login_required(f):
    """
    Decorate routes to require login.

    https://flask.palletsprojects.com/en/latest/patterns/viewdecorators/
    """

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("user_id") is None:
            return redirect("/login")
        return f(*args, **kwargs)

    return decorated_function


def generate_thumbnails(video_dir,thumbnail_dir,resize=2):
    '''
    Provide paths and value to resize.
    Generate thumbnails for videos in a directory
    '''

    paths = {}
    duration = {}
    for video_file in os.listdir(video_dir):
        if video_file.endswith('.mov'):
            video_path = os.path.join(video_dir, video_file)
            thumbnail_path, video_duration = extract_thumbnail(video_path, thumbnail_dir)
            resize_image(thumbnail_path, resize)
            thumbnail_name = video_file.replace('.mov', '.jpeg')
            paths[thumbnail_name] = thumbnail_path
            duration[thumbnail_name] = video_duration

    return paths, duration

def resize_image(image_path, val):
    """Resize image by a given factor."""

    with Image.open(image_path) as img:
        # Resize image to half its size
        img = img.resize((img.width // val, img.height // val), Image.LANCZOS)
        img.save(image_path)



def extract_thumbnail(video_path, thumbnail_dir):
    """
    Extract thumbnail from a video.
    Function will get back duration and thumbnail_path
    """

    video = VideoFileClip(video_path)
    thumbnail_path = os.path.join(thumbnail_dir, os.path.basename(video_path).replace('.mov', '.jpeg'))
    video.save_frame(thumbnail_path, t=video.duration / 2)
    total_frames = math.ceil(video.duration * video.fps) # pass total_frames back (gonna use for excel)

    return thumbnail_path, total_frames




def create_excel(image_folder,data):
    '''
    Create an Excel file with image data.
    '''

    excel_path = os.path.join(image_folder, "Excel.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active # Current worksheet

    columns = ['A','B','C','D','E','F','G']
    headers = ["№","SHOT PREVIEW","SHOT NAME","DURATION","VFX SCOPE OF WORK DESCRIPTION","DAYS","COST"]

    for col,header in zip(columns,headers):
        ws[col + '1'] = header
        ws.column_dimensions[col].width = len(header) + 5


    row_index = 2
    for image_file in os.listdir(image_folder):
        if image_file.endswith('.jpeg'):
            # add image
            image_path = os.path.join(image_folder, image_file)
            img = OImage(image_path)

            ws.row_dimensions[row_index].height = img.height
            ws.column_dimensions['B'].width = img.width/6
            ws.add_image(img,'B' + str(row_index))


            # add duration
            ws['D' + str(row_index)] = data[image_file]


            # add shot name
            ws['C' + str(row_index)] = Path(image_file).stem
            ws.column_dimensions['C'].width = len(Path(image_file).stem)


            # add shot number
            ws['A' + str(row_index)] = (row_index - 1)

            row_index += 1

    wb.save(excel_path)


